import inspect
import logging
import softest
from openpyxl import Workbook, load_workbook

class Utils(softest.TestCase):
    def assert_list_item_text(self, lst, value):
        for items in lst:
            print("The text is: " + items.text)
            self.soft_assert(self.assertEqual, items.text, value)
            if items.text == value:
                print("test passed")
            else:
                print("test failed")
        self.assert_all()


    def custom_logger(self,log_level=logging.DEBUG):
        try:
            # Set class/method name from where it's called
            logger_name = inspect.stack()[1][3]
            # Create logger
            logger = logging.getLogger(logger_name)
            logger.setLevel(log_level)
            # Create file handler and set the log level
            fh = logging.FileHandler("automation.log", mode='a')
            # Create formatter - how you want your logs to be formatted
            formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s : %(message)s',
                                          datefmt='%m/%d/%Y %I:%M:%S %p')
            # Add formatter to file handler
            fh.setFormatter(formatter)
            # Add file handler to logger
            logger.addHandler(fh)
            return logger

        except Exception as e:
            # Handle exceptions and log the error
            logging.error(f"Error setting up logger: {str(e)}")
            return None
    @staticmethod
    def read_excel_data(file_name,sheet):
        try:
            # Initialize empty data_list
            data_lst = []
            # Load workbook
            wb = load_workbook(filename=file_name)
            sh = wb[sheet]  # Active Sheet
            # Row Count
            row_ct = sh.max_row
            # Col Count
            col_ct = sh.max_column

            for i in range(2, row_ct + 1):
                # Initialize a variable to store the row data
                row_dt = []
                for j in range(1, col_ct + 1):
                    row_dt.append(sh.cell(row=i, column=j).value)
                data_lst.append(row_dt)  # Append row data outside the inner loop
            return data_lst

        except FileNotFoundError:
            logging.error(f"Error: The file '{file_name}' was not found.")
            return []

    @staticmethod
    def verify_text(actual_text, expected_text, logger, success_message="Text verified successfully"):
        """
        Verifies if the actual text matches the expected text.

        Args:
            actual_text (str): The actual text retrieved from the web page.
            expected_text (str): The expected text to compare against.
            logger: The logging object to log messages.
            success_message (str): Optional success message for logging.

        Returns:
            None: Raises an assertion error if the texts do not match.
        """
        assert actual_text == expected_text, f"Text does not match! Expected: '{expected_text}', Got: '{actual_text}'"
        logger.info(success_message)


